import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO, StringIO
from datetime import date, timedelta
import csv as _csv
import json
import httpx
import logging
import re
from services.rate_matcher import find_rate, build_rates_lookup, CURRENCY_MAP
from config import settings

logger = logging.getLogger(__name__)
PRODUCT_FOOTER_TEXT = "Xchange Book | A product of Patience AI | https://patienceai.in"
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


# ──────────────────────────────────────────────────────────────────────────────
# Robust CSV reading — real-world expense CSVs often have a title/metadata row
# above the header, trailing commas, or ragged rows. pandas' C parser locks the
# column count to the FIRST line and raises "Expected N fields" on such files.
# We read via the stdlib csv reader (ragged-tolerant), pad every row to the
# widest row, and build the DataFrame ourselves so header detection stays aligned.
# ──────────────────────────────────────────────────────────────────────────────
def _decode_csv(file_bytes: bytes) -> str:
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return file_bytes.decode(enc)
        except UnicodeDecodeError:
            continue
    return file_bytes.decode("utf-8", errors="replace")


def _sniff_delimiter(text: str) -> str:
    sample = text[:8192]
    try:
        return _csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
    except Exception:
        first_line = next((ln for ln in sample.splitlines() if ln.strip()), "")
        counts = {d: first_line.count(d) for d in (",", ";", "\t", "|")}
        best = max(counts, key=counts.get)
        return best if counts[best] > 0 else ","


def _read_csv_robust(file_bytes: bytes, header):
    """Ragged-tolerant CSV read.
    header=None -> full matrix with a RangeIndex header (like read_csv(header=None)).
    header=<int> -> use that row as the header, rows below as data.
    Rows are kept 1:1 (only padded) so a header index computed on the header=None
    matrix stays valid here."""
    text = _decode_csv(file_bytes)
    delimiter = _sniff_delimiter(text)
    rows = list(_csv.reader(StringIO(text), delimiter=delimiter))
    max_w = max((len(r) for r in rows), default=0)
    if max_w == 0:
        return pd.DataFrame()
    norm = [r + [None] * (max_w - len(r)) for r in rows]
    full = pd.DataFrame(norm)
    if header is None:
        return full
    header_vals = [str(c).strip() if c is not None else "" for c in full.iloc[header].tolist()]
    body = full.iloc[header + 1:].copy()
    body.columns = header_vals
    return body.reset_index(drop=True)


def _detect_header_row(df_raw: pd.DataFrame) -> int:
    """The first row with >=3 non-empty string cells is the header. Real-world
    files often carry a title/blank row above it — skipping to this row is what
    keeps columns named correctly instead of 'Unnamed: N'."""
    for i, row in df_raw.iterrows():
        cells = [str(v).strip() for v in row.dropna() if str(v).strip()]
        if len(cells) >= 3:
            return int(i)
    return 0


async def detect_columns_with_ai(columns: list, sample_rows: list) -> dict:
    prompt = f"""Given these spreadsheet columns and sample data, identify:
- date_col: column containing transaction/expense date
- currency_col: column containing 3-letter currency code like USD, INR, EUR
- amount_col: column containing amount/value
- ref_rate_col: column for reference/exchange rate (may be empty)

Columns: {columns}
Sample data (first 5 rows): {json.dumps(sample_rows, default=str)}

Return ONLY valid JSON:
{{"date_col": "column_name_or_null", "currency_col": "column_name_or_null", "amount_col": "column_name_or_null", "ref_rate_col": "column_name_or_null"}}"""

    try:
        async with httpx.AsyncClient(timeout=30) as client:
            resp = await client.post(
                "https://openrouter.ai/api/v1/chat/completions",
                headers={"Authorization": f"Bearer {settings.OPENROUTER_API_KEY}"},
                json={
                    "model": settings.OPENROUTER_MODEL,
                    "messages": [{"role": "user", "content": prompt}],
                    "max_tokens": 500,
                }
            )
            result = resp.json()
            content = result["choices"][0]["message"]["content"].strip()
            json_match = re.search(r'\{[^{}]+\}', content, re.DOTALL)
            if json_match:
                return json.loads(json_match.group())
    except Exception as e:
        logger.error(f"AI column detection failed: {e}")

    return {"date_col": None, "currency_col": None, "amount_col": None, "ref_rate_col": None}

def detect_columns(df: pd.DataFrame) -> dict:
    cols_lower = {str(c).strip().lower(): str(c).strip() for c in df.columns}
    result = {}

    for key in ["expense date", "date", "transaction date", "invoice date", "exp date"]:
        if key in cols_lower:
            result["date_col"] = cols_lower[key]
            break

    for key in ["currency", "currency type", "curr", "ccy"]:
        if key in cols_lower:
            result["currency_col"] = cols_lower[key]
            break

    for key in ["amount", "original amount", "value", "amt"]:
        if key in cols_lower:
            result["amount_col"] = cols_lower[key]
            break

    for key in ["ref rate", "ref_rate", "reference rate", "exchange rate", "fx rate", "refrate"]:
        if key in cols_lower:
            result["ref_rate_col"] = cols_lower[key]
            break

    return result


def _parse_date(raw_date) -> date | None:
    if pd.isna(raw_date) if hasattr(raw_date, '__class__') else raw_date is None:
        return None
    try:
        if isinstance(raw_date, (int, float)):
            # Excel serial date (days since 1899-12-30)
            return (pd.Timestamp("1899-12-30") + pd.Timedelta(days=int(raw_date))).date()
        if hasattr(raw_date, 'date'):
            return raw_date.date()
        return pd.to_datetime(str(raw_date)).date()
    except Exception:
        return None


async def process_expense_file(
    file_bytes: bytes,
    filename: str,
    rates_rows: list,
    progress_callback=None
) -> tuple[bytes, dict]:
    """
    Process an expense file filling in FBIL reference rates.
    Returns (processed_bytes, stats_dict)
    """
    is_xlsx = filename.lower().endswith(".xlsx")
    rates_lookup = build_rates_lookup(rates_rows)

    if progress_callback:
        await progress_callback(10, "reading", "Reading uploaded file...")

    # ── Step 1: Detect header row ──────────────────────────────────────────
    if is_xlsx:
        df_raw = pd.read_excel(BytesIO(file_bytes), header=None)
    else:
        df_raw = _read_csv_robust(file_bytes, header=None)

    header_row_idx = _detect_header_row(df_raw)

    logger.info(f"Detected header at row {header_row_idx}")

    if is_xlsx:
        df = pd.read_excel(BytesIO(file_bytes), header=header_row_idx)
    else:
        df = _read_csv_robust(file_bytes, header=header_row_idx)

    df.columns = [str(c).strip() for c in df.columns]
    # Drop completely empty rows
    df = df.dropna(how="all").reset_index(drop=True)

    if progress_callback:
        await progress_callback(20, "detecting", "Detecting column structure...")

    # ── Step 2: Detect columns ─────────────────────────────────────────────
    col_map = detect_columns(df)

    if not col_map.get("date_col") or not col_map.get("currency_col"):
        sample = df.head(5).to_dict(orient="records")
        logger.info("Heuristic detection failed, trying fallback AI...")
        ai_result2 = await detect_columns_with_ai(list(df.columns), sample)
        for k, v in ai_result2.items():
            if v and str(v) in df.columns and k not in col_map:
                col_map[k] = str(v)

    date_col = col_map.get("date_col")
    currency_col = col_map.get("currency_col")
    ref_rate_col = col_map.get("ref_rate_col") or "Ref Rate"

    logger.info(f"Columns: date={date_col}, currency={currency_col}, ref_rate={ref_rate_col}")

    if not date_col or not currency_col:
        raise ValueError(f"Cannot find date/currency columns. Found: {list(df.columns)}")

    # Add ref rate column if missing
    if ref_rate_col not in df.columns:
        df[ref_rate_col] = None

    if progress_callback:
        await progress_callback(35, "matching", "Matching FBIL rates to expense rows...")

    # ── Step 3: Match rates ────────────────────────────────────────────────
    total_rows = 0
    matched_rows = 0
    unmatched_rows = 0
    skipped_no_date = 0
    unmatched_dates = []

    for idx in range(len(df)):
        row = df.iloc[idx]
        currency_val = str(row.get(currency_col, "")).strip().upper()

        # Skip domestic INR and empty rows
        if not currency_val or currency_val in ("INR", "NAN", "", "NONE", "NAT"):
            continue

        parsed_date = _parse_date(row.get(date_col))

        # A row with a foreign currency but no usable date cannot be matched to a
        # rate (there is nothing to look up). Count it as "skipped — no date"
        # rather than "unmatched", so the match rate reflects only rows that
        # actually carry a date. We never guess/forward-fill dates — that could
        # attach a wrong rate to a financial figure.
        if parsed_date is None:
            skipped_no_date += 1
            continue

        total_rows += 1

        # Emit progress every 30 rows
        if total_rows % 30 == 0 and progress_callback:
            pct = 35 + min(40, int((idx / max(len(df), 1)) * 40))
            await progress_callback(pct, "matching", f"Matching rates: {matched_rows}/{total_rows} rows...")

        rate = find_rate(parsed_date, currency_val, rates_lookup)
        if rate is not None:
            df.at[idx, ref_rate_col] = rate
            matched_rows += 1
        else:
            unmatched_rows += 1
            unmatched_dates.append(str(parsed_date))

    if progress_callback:
        await progress_callback(78, "writing", "Writing processed file preserving formatting...")

    # ── Step 4: Write output ───────────────────────────────────────────────
    if is_xlsx:
        processed_bytes = _write_xlsx(file_bytes, df, header_row_idx, ref_rate_col)
    else:
        out = BytesIO()
        df.to_csv(out, index=False)
        out.write(f"\n\n{PRODUCT_FOOTER_TEXT}\n".encode("utf-8"))
        processed_bytes = out.getvalue()

    if progress_callback:
        await progress_callback(92, "uploading", "Uploading to cloud storage...")

    stats = {
        "total_rows": total_rows,
        "matched_rows": matched_rows,
        "unmatched_rows": unmatched_rows,
        "skipped_no_date": skipped_no_date,
        "unmatched_dates": sorted(set(unmatched_dates))[:30],
        "date_col": date_col,
        "currency_col": currency_col,
        "ref_rate_col": ref_rate_col,
        "match_rate_pct": round(matched_rows / max(total_rows, 1) * 100, 1),
    }

    return processed_bytes, stats


def _write_xlsx(original_bytes: bytes, df: pd.DataFrame, header_row_idx: int, ref_rate_col: str) -> bytes:
    """Write processed XLSX preserving original formatting, filling ref rate column."""
    try:
        wb = load_workbook(BytesIO(original_bytes))
        ws = wb.active

        # openpyxl is 1-indexed; header is at header_row_idx+1
        header_ws_row = header_row_idx + 1

        # Find the ref rate column index in the worksheet
        ref_col_idx = None
        for cell in ws[header_ws_row]:
            if cell.value and str(cell.value).strip() == ref_rate_col:
                ref_col_idx = cell.column
                break

        if ref_col_idx is None:
            # Append as new column
            max_col = ws.max_column
            ref_col_idx = max_col + 1
            header_cell = ws.cell(row=header_ws_row, column=ref_col_idx, value=ref_rate_col)
            header_cell.font = Font(bold=True, color="000000")
            header_cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        # Map df index to worksheet row
        # df row 0 → ws row header_ws_row+1
        for df_idx in range(len(df)):
            ws_row = header_ws_row + 1 + df_idx
            if ws_row > ws.max_row:
                break

            rate_val = df.iloc[df_idx].get(ref_rate_col)
            if rate_val is not None and str(rate_val) not in ("nan", "None", ""):
                try:
                    cell = ws.cell(row=ws_row, column=ref_col_idx, value=float(rate_val))
                    cell.number_format = '0.0000'
                except Exception:
                    pass

        out = BytesIO()
        footer_row = header_ws_row + 1 + len(df) + 2
        footer_cell = ws.cell(row=footer_row, column=1, value="Xchange Book")
        footer_cell.font = Font(bold=True, color="FFFFFF")
        footer_cell.fill = PatternFill(start_color="1A2035", end_color="1A2035", fill_type="solid")
        footer_cell.alignment = Alignment(horizontal="left")

        company_cell = ws.cell(row=footer_row + 1, column=1, value="A product of Patience AI")
        company_cell.font = Font(color="0B6E4F", bold=True)
        link_cell = ws.cell(row=footer_row + 2, column=1, value="https://patienceai.in")
        link_cell.hyperlink = "https://patienceai.in"
        link_cell.style = "Hyperlink"

        wb.save(out)
        return out.getvalue()

    except Exception as e:
        logger.error(f"openpyxl write error: {e}, falling back to pandas")
        out = BytesIO()
        with pd.ExcelWriter(out, engine="openpyxl") as writer:
            df.to_excel(writer, index=False)
        return out.getvalue()


# ──────────────────────────────────────────────────────────────────────────────
# Multi-format export — a processed file (stored in its original xlsx/csv form)
# can be downloaded as xlsx, csv, or pdf. Same-format requests return the stored
# bytes untouched (preserves the xlsx formatting/footer); other formats convert.
# ──────────────────────────────────────────────────────────────────────────────
def _processed_to_df(file_bytes: bytes, stored_filename: str) -> pd.DataFrame:
    # Read raw (no header) first, detect the real header row — the processed
    # xlsx keeps any title/blank rows the original had above the header, so a
    # naive header=0 read would label every column "Unnamed: N".
    if stored_filename.lower().endswith(".xlsx"):
        raw = pd.read_excel(BytesIO(file_bytes), header=None)
        hdr = _detect_header_row(raw)
        df = pd.read_excel(BytesIO(file_bytes), header=hdr)
    else:
        raw = _read_csv_robust(file_bytes, header=None)
        hdr = _detect_header_row(raw)
        df = _read_csv_robust(file_bytes, header=hdr)

    df.columns = [str(c).strip() for c in df.columns]

    # Strip trailing branded footer / fully-empty rows added at write time
    df = df.dropna(how="all")
    footer_mask = df.apply(
        lambda r: r.astype(str).str.contains("Xchange Book|Patience AI|patienceai.in", case=False, na=False).any(),
        axis=1,
    )
    df = df[~footer_mask].reset_index(drop=True)

    # Drop columns that have no header AND no data (e.g. a blank leading column
    # that pandas names "Unnamed: 0"); blank the header of unnamed-but-populated
    # columns so the export never shows "Unnamed: N".
    keep, rename = [], {}
    for c in df.columns:
        is_unnamed = str(c).startswith("Unnamed") or str(c).strip() == "" or str(c).lower() == "nan"
        if is_unnamed and df[c].isna().all():
            continue
        keep.append(c)
        if is_unnamed:
            rename[c] = ""
    df = df[keep].rename(columns=rename)
    return df


def convert_processed(file_bytes: bytes, stored_filename: str, target_format: str):
    """Return (bytes, media_type, extension) for the requested download format."""
    target = (target_format or "").strip().lower()
    stored_ext = stored_filename.lower().rsplit(".", 1)[-1] if "." in stored_filename else "xlsx"

    if target in ("", stored_ext):
        media = XLSX_MIME if stored_ext == "xlsx" else "text/csv"
        return file_bytes, media, stored_ext

    if target in ("xlsx", "excel"):
        if stored_ext == "xlsx":
            return file_bytes, XLSX_MIME, "xlsx"
        df = _processed_to_df(file_bytes, stored_filename)
        out = BytesIO()
        with pd.ExcelWriter(out, engine="openpyxl") as writer:
            df.to_excel(writer, index=False)
        return out.getvalue(), XLSX_MIME, "xlsx"

    if target == "csv":
        if stored_ext == "csv":
            return file_bytes, "text/csv", "csv"
        df = _processed_to_df(file_bytes, stored_filename)
        out = BytesIO()
        df.to_csv(out, index=False)
        out.write(f"\n\n{PRODUCT_FOOTER_TEXT}\n".encode("utf-8"))
        return out.getvalue(), "text/csv", "csv"

    if target == "pdf":
        df = _processed_to_df(file_bytes, stored_filename)
        return _df_to_pdf(df), "application/pdf", "pdf"

    # Unknown format → stored bytes untouched
    media = XLSX_MIME if stored_ext == "xlsx" else "text/csv"
    return file_bytes, media, stored_ext


def _df_to_pdf(df: pd.DataFrame) -> bytes:
    """Render a DataFrame to a branded landscape PDF table (reportlab imported
    lazily so it never costs memory until a PDF is actually requested)."""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    MAX_COLS, MAX_ROWS = 12, 3000
    cols = [str(c) for c in df.columns][:MAX_COLS]
    truncated_cols = len(df.columns) > MAX_COLS
    body_df = df.head(MAX_ROWS)
    truncated_rows = len(df) > MAX_ROWS

    styles = getSampleStyleSheet()
    head_style = ParagraphStyle("h", parent=styles["BodyText"], fontSize=7, leading=9,
                                textColor=colors.white, fontName="Helvetica-Bold")
    cell_style = ParagraphStyle("c", parent=styles["BodyText"], fontSize=6.8, leading=8.4)

    def _fmt(v):
        if pd.isna(v):
            return ""
        s = str(v)
        return (s[:60] + "…") if len(s) > 61 else s

    data = [[Paragraph(_fmt(c), head_style) for c in cols]]
    for _, row in body_df.iterrows():
        data.append([Paragraph(_fmt(row[c]), cell_style) for c in cols])

    page_w = landscape(A4)[0] - 20 * mm
    col_w = page_w / max(len(cols), 1)
    table = Table(data, colWidths=[col_w] * len(cols), repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1A2035")),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D0D5DD")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F4F6FB")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))

    title_style = ParagraphStyle("t", parent=styles["Title"], fontSize=15, textColor=colors.HexColor("#1A2035"))
    note_style = ParagraphStyle("n", parent=styles["Normal"], fontSize=7, textColor=colors.HexColor("#667085"))
    foot_style = ParagraphStyle("f", parent=styles["Normal"], fontSize=7.5, textColor=colors.HexColor("#0B6E4F"))

    story = [Paragraph("Xchange Book — Processed Expense Report", title_style), Spacer(1, 6), table, Spacer(1, 8)]
    if truncated_cols or truncated_rows:
        parts = []
        if truncated_rows:
            parts.append(f"first {MAX_ROWS} rows")
        if truncated_cols:
            parts.append(f"first {MAX_COLS} columns")
        story.append(Paragraph("Note: PDF shows " + " and ".join(parts) + ". Use Excel/CSV for the full data.", note_style))
        story.append(Spacer(1, 4))
    story.append(Paragraph(PRODUCT_FOOTER_TEXT, foot_style))

    buf = BytesIO()
    SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=10 * mm, rightMargin=10 * mm,
                      topMargin=12 * mm, bottomMargin=12 * mm).build(story)
    return buf.getvalue()
